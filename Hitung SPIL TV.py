import streamlit as st
import zipfile
import tempfile
import os
import io
import shutil
import pandas as pd
import csv
import sys

st.set_page_config(page_title="Rekap SPIL TV", layout="wide")
st.title("Calculator Rekap Kehadiran SPIL TV")

st.markdown("""
Upload sebuah file ZIP yang berisi Excel (.xlsx/.xls) dan/atau CSV.  
""")

uploaded = st.file_uploader("Upload file ZIP", type=["zip"])

st.info("Dependencies: `xlsx2csv` direkomendasikan (pip install xlsx2csv). App akan mencoba fallback jika tidak tersedia.")

if uploaded is not None:
    if st.button("Proses"):
        with st.spinner("Memproses..."):
            tmpd = tempfile.mkdtemp()
            try:
                zip_bytes = uploaded.read()
                zip_path = os.path.join(tmpd, "input.zip")
                with open(zip_path, "wb") as f:
                    f.write(zip_bytes)

                extracted_dir = os.path.join(tmpd, "extracted")
                os.makedirs(extracted_dir, exist_ok=True)
                with zipfile.ZipFile(zip_path, "r") as z:
                    z.extractall(extracted_dir)

                excel_paths = []
                csv_paths = []

                for root, _, files in os.walk(extracted_dir):
                    for fn in files:
                        lower = fn.lower()
                        full = os.path.join(root, fn)
                        if lower.endswith(".xlsx") or lower.endswith(".xls"):
                            excel_paths.append(full)
                        elif lower.endswith(".csv"):
                            csv_paths.append(full)

                st.write(f"Menemukan **{len(excel_paths)}** file Excel dan **{len(csv_paths)}** file CSV di dalam ZIP.")

                # Convert XLSX→CSV
                def convert_xlsx_to_csv(src_path, out_path):
                    try:
                        from xlsx2csv import Xlsx2csv
                        Xlsx2csv(src_path, outputencoding="utf-8").convert(out_path)
                        return True, None
                    except Exception as e_x:
                        try:
                            import openpyxl
                            wb = openpyxl.load_workbook(src_path, data_only=True, read_only=True)
                            sheet_name = "Result" if "Result" in wb.sheetnames else wb.sheetnames[0]
                            ws = wb[sheet_name]
                            with open(out_path, "w", encoding="utf-8", newline="") as f:
                                writer = csv.writer(f)
                                for row in ws.iter_rows(values_only=True):
                                    writer.writerow([("" if v is None else v) for v in row])
                            return True, None
                        except Exception as e2:
                            return False, f"xlsx2csv error: {e_x}; openpyxl fallback error: {e2}"

                csv_out_dir = os.path.join(tmpd, "csvs")
                os.makedirs(csv_out_dir, exist_ok=True)
                converted_csv_paths = []
                failed_conversions = []

                for xp in excel_paths:
                    base = os.path.basename(xp)
                    name_no_ext = os.path.splitext(base)[0]
                    outname = name_no_ext + ".csv"
                    outpath = os.path.join(csv_out_dir, outname)
                    ok, err = convert_xlsx_to_csv(xp, outpath)
                    if ok:
                        converted_csv_paths.append(outpath)
                    else:
                        failed_conversions.append((xp, err))

                for cp in csv_paths:
                    dest = os.path.join(csv_out_dir, os.path.basename(cp))
                    try:
                        shutil.copy(cp, dest)
                        converted_csv_paths.append(dest)
                    except:
                        try:
                            with open(cp, "rb") as r:
                                data = r.read()
                            with open(dest, "wb") as w:
                                w.write(data)
                            converted_csv_paths.append(dest)
                        except:
                            pass

                # ZIP all CSVs
                csv_zip_bytes = None
                if converted_csv_paths:
                    mem_zip = io.BytesIO()
                    with zipfile.ZipFile(mem_zip, "w", zipfile.ZIP_DEFLATED) as zout:
                        for p in converted_csv_paths:
                            try:
                                zout.write(p, os.path.basename(p))
                            except:
                                with open(p, "rb") as rf:
                                    zout.writestr(os.path.basename(p), rf.read())
                    mem_zip.seek(0)
                    csv_zip_bytes = mem_zip.read()

                # === REKAP PROSES ===
                rekap_rows = []
                if converted_csv_paths:
                    for p in converted_csv_paths:
                        fname = os.path.basename(p)

                        # Ambil raw baris 2 kolom 1 (judul)
                        title_raw = ""
                        try:
                            with open(p, "r", encoding="utf-8", errors="replace") as f:
                                lines = f.readlines()
                        except:
                            with open(p, "r", encoding="latin1", errors="replace") as f:
                                lines = f.readlines()

                        if len(lines) > 1:
                            # Pecah per koma dan ambil KOLUM PERTAMA SAJA
                            first_row_split = lines[1].split(",")
                            title_raw = first_row_split[0].strip()
                        else:
                            title_raw = ""

                        # Pisahkan JUDUL dan TANGGAL
                        judul_clean = ""
                        tanggal_clean = ""

                        raw = title_raw.strip()

                        # 1) Jika format lama: "Judul / Tanggal"
                        if " / " in raw:
                            parts = raw.split(" / ", 1)
                            judul_clean = parts[0].strip()
                            tanggal_clean = parts[1].strip()

                        else:
                            # 2) Format baru: "Judul - 28 November 2025"
                            if " - " in raw:
                                parts = raw.rsplit(" - ", 1)  # pecah dari belakang
                                left, right = parts[0].strip(), parts[1].strip()

                                # deteksi apakah bagian kanan adalah tanggal valid
                                bulan_list = [
                                    "januari","februari","maret","april","mei","juni","juli",
                                    "agustus","september","oktober","november","desember"
                                ]

                                lower_right = right.lower()

                                if any(b in lower_right for b in bulan_list):
                                    judul_clean = left
                                    tanggal_clean = right
                                else:
                                    # fallback kalau ternyata bukan tanggal
                                    judul_clean = raw
                                    tanggal_clean = ""
                            else:
                                # fallback terakhir
                                judul_clean = raw
                                tanggal_clean = ""

                        # Hitung jumlah peserta
                        peserta_count = 0
                        try:
                            df = None
                            try:
                                df = pd.read_csv(p, skiprows=2, dtype=str, on_bad_lines="skip", encoding="utf-8")
                            except:
                                df = pd.read_csv(p, skiprows=2, dtype=str, on_bad_lines="skip", encoding="latin1")

                            if df is not None and df.shape[0] > 0:
                                first_row_vals = df.iloc[0].astype(str).str.cat(sep=" ").lower()
                                if ("nik" in first_row_vals) or ("no" in first_row_vals and "nik" in first_row_vals):
                                    df.columns = df.iloc[0]
                                    df = df.iloc[1:].reset_index(drop=True)

                                peserta_count = df[df.iloc[:, 0].notna()].shape[0]
                        except:
                            peserta_count = 0

                        rekap_rows.append({
                            "Nama File": fname,
                            "Judul": judul_clean,
                            "Tanggal": tanggal_clean,
                            "Jumlah Peserta": peserta_count
                        })

                # OUTPUT
                cols1, cols2 = st.columns(2)

                if csv_zip_bytes:
                    cols1.success(f"{len(converted_csv_paths)} CSV tersedia.")
                    cols1.download_button(
                        "📥 Download ZIP berisi semua CSV",
                        csv_zip_bytes,
                        "all_csvs.zip",
                        "application/zip"
                    )
                else:
                    cols1.info("Tidak ada CSV untuk di-download.")

                if rekap_rows:
                    df_rekap = pd.DataFrame(rekap_rows)
                    st.dataframe(df_rekap)

                    buf = io.StringIO()
                    df_rekap.to_csv(buf, index=False)
                    cols2.success("Rekap siap.")
                    cols2.download_button(
                        "📥 Download Rekap (CSV)",
                        buf.getvalue().encode("utf-8"),
                        "rekap_kehadiran.csv",
                        "text/csv"
                    )
                else:
                    cols2.info("Tidak ada data untuk direkap.")

                if failed_conversions:
                    st.warning("Beberapa file Excel gagal dikonversi:")
                    for p, err in failed_conversions:
                        st.write(f"- {os.path.basename(p)} : {err}")

            finally:
                try:
                    shutil.rmtree(tmpd)
                except:
                    pass

        st.success("Selesai.")
